#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from llm_client import DeepSeekClient, LLMRequestError


NOT_SMELL = "Not a smell"

TAXONOMY: Dict[str, str] = {
    "Misleading": "Comment says something that's inconsistent or incorrect relative to what the code does.",
    "Obvious": "Comments that merely restate the code without adding useful information.",
    "Commented out code": "A code piece that is commented out.",
    "Irrelevant": "Comments that do not intend to explain the code.",
    "Task": "Comments explaining the work that could/should be done in future or was already completed.",
    "Too much info": "Comments that are overly detailed or verbose, usually stating implementation-level trivia.",
    "Beautification": "Comments that aim to distinguish the parts of the code.",
    "Nonlocal info": "Comments that provide systemwide information or mention code that is not near.",
    "Vague": "Comments that are not clearly understandable.",
    "Attribution": "Comments that give information about who wrote the code.",
    "Not a smell": "Comments that enhance code comprehension, provide context or explain complex logic.",
}

LABEL_TO_ACTION = {
    "Misleading": "Rewrite",
    "Obvious": "Remove",
    "Vague": "Rewrite",
    "Too much info": "Summarize",
    "Task": "Remove",
    "Commented out code": "Remove",
    "Irrelevant": "Remove",
    "Nonlocal info": "Move",
    "Beautification": "Remove",
    "Attribution": "Remove",
    "Not a smell": "No Change",
}


@dataclass
class RepairOutput:
    sample_id: str
    predicted_label: str
    detection_confidence: float
    repair_action: str
    original_comment: str
    repaired_comment: Optional[str]
    repair_status: str
    reason_short: str
    route_to: str
    model_source: str


@dataclass
class RepairTrace:
    sample_id: str
    language: str
    code: str
    original_comment: str
    predicted_label: str
    detection_confidence: float
    first_pass_raw: Optional[Dict[str, Any]]
    second_pass_raw: Optional[Dict[str, Any]]
    repair: RepairOutput



def resolve_dataset_path(path: Optional[Path] = None) -> Path:
    if path is not None:
        return path
    p1 = Path("Data") / "Dataset.xlsx"
    p2 = Path("data") / "Dataset.xlsx"
    if p1.exists():
        return p1
    return p2


def norm(s: Optional[str]) -> str:
    return "" if s is None else str(s).strip()


def canon_label(label: str) -> str:
    raw = norm(label)
    low = raw.lower().replace("-", " ")
    aliases = {
        "not a smell": "Not a smell",
        "too much information": "Too much info",
        "too much info": "Too much info",
        "commented out code": "Commented out code",
        "commented out code ": "Commented out code",
        "commented out code": "Commented out code",
        "nonlocal": "Nonlocal info",
        "nonlocal info": "Nonlocal info",
        "non local": "Nonlocal info",
        "non-local": "Nonlocal info",
        "attribution": "Attribution",
    }
    return aliases.get(low, raw)


def find_headers(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        mapping[str(v).strip().lower()] = col
    return mapping


def get_cell(ws, row: int, headers: Dict[str, int], names: Sequence[str]) -> str:
    for name in names:
        idx = headers.get(name.lower())
        if idx is None:
            continue
        v = ws.cell(row=row, column=idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def load_dataset_map(path: Path, limit: Optional[int] = None) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = find_headers(ws)

    out: Dict[str, Dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        sid = get_cell(ws, row, headers, ["sample_id", "id"])
        if not sid:
            sid = str(row - 1)
        out[sid] = {
            "sample_id": sid,
            "language": get_cell(ws, row, headers, ["language", "lang"]),
            "code": get_cell(ws, row, headers, ["code"]),
            "comment": get_cell(ws, row, headers, ["comment"]),
            "gold_label": get_cell(ws, row, headers, ["gold_label"]),
            "repair_action": get_cell(ws, row, headers, ["repair_action"]),
            "gold_repair": get_cell(ws, row, headers, ["gold_repair"]),
        }
        if limit is not None and len(out) >= limit:
            break

    return out


def load_detection_trace(path: Path, limit: Optional[int] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            rows.append(json.loads(line))
            if limit is not None and len(rows) >= limit:
                break
    return rows


class RepairAgent:
    def __init__(
        self,
        temperature: float = 0.2,
    ) -> None:
        self.client = DeepSeekClient()
        self.temperature = temperature

    def _rewrite_or_summarize(self, action: str, language: str, code: str, comment: str, label: str) -> Dict[str, Any]:
        system_prompt = (
            "You are a code comment repair model. "
            "Return JSON only and keep comment faithful to code."
        )
        user_prompt = f"""
Task: Repair inline comment.
Required action: {action}
Predicted smell label: {label}
Language: {language}

Rules:
- Do not change code.
- Output only one repaired comment string.
- Keep concise and precise.
- If action=Summarize, shorten and preserve intent.
- If action=Rewrite, rewrite clearly and faithfully.

Label taxonomy:
{json.dumps(TAXONOMY, ensure_ascii=False)}

Return JSON:
{{
  "repaired_comment": "...",
  "reason_short": "..."
}}

Code:
{code}

Original comment:
{comment}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=self.temperature, max_tokens=500)

    def repair_one(self, record: Dict[str, Any], dataset_row: Dict[str, str]) -> Tuple[RepairOutput, Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
        det = record.get("detection", {})
        sid = str(record.get("sample_id", dataset_row.get("sample_id", "")))
        language = record.get("language") or dataset_row.get("language", "")
        code = record.get("code") or dataset_row.get("code", "")
        comment = record.get("comment") or dataset_row.get("comment", "")

        label = canon_label(norm(det.get("label") or "Not a smell"))
        conf = float(det.get("confidence", 0.0))
        action = LABEL_TO_ACTION.get(label, "No Change")
        if label == "Obvious":
            comment_low = norm(comment).lower()
            has_intent_signal = any(
                k in comment_low
                for k in ["because", "so that", "to avoid", "to ensure", "so ", "for ", "under load", "risk", "constraint"]
            )
            short_and_plain = len(comment_low.split()) <= 7
            action = "Remove" if (conf >= 0.90 and short_and_plain and not has_intent_signal) else "Rewrite"

        first_raw: Optional[Dict[str, Any]] = None
        second_raw: Optional[Dict[str, Any]] = None

        repaired_comment: Optional[str]
        reason_short = ""
        route_to = "evaluation"

        if label == NOT_SMELL or action == "No Change":
            repaired_comment = None
            status = "skipped_clean"
            reason_short = "Predicted clean; no repair attempted."
        elif action == "Remove":
            repaired_comment = ""
            status = "removed"
            reason_short = "Removed by policy for this smell label."
        elif action == "Move":
            repaired_comment = ""
            status = "moved"
            reason_short = "Moved nonlocal note out of inline comment."
        elif action in {"Rewrite", "Summarize"}:
            first_raw = self._rewrite_or_summarize(action, str(language), str(code), str(comment), label)
            repaired_comment = norm(first_raw.get("repaired_comment"))
            reason_short = norm(first_raw.get("reason_short"))
            if not repaired_comment:
                repaired_comment = comment
                reason_short = reason_short or "Model returned empty comment; fallback to original."
                status = "fallback_original"
            else:
                status = "repaired"
        else:
            repaired_comment = comment
            status = "no_change"
            reason_short = "Unsupported action; keep original."

        out = RepairOutput(
            sample_id=sid,
            predicted_label=label,
            detection_confidence=round(max(0.0, min(1.0, conf)), 4),
            repair_action=action,
            original_comment=str(comment),
            repaired_comment=repaired_comment,
            repair_status=status,
            reason_short=reason_short[:220],
            route_to=route_to,
            model_source=self.client.model_source,
        )
        return out, first_raw, second_raw


def write_outputs(results: List[RepairOutput], jsonl_path: Path, csv_path: Path) -> None:
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    with jsonl_path.open("w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(asdict(r), ensure_ascii=False) + "\n")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "sample_id",
            "predicted_label",
            "detection_confidence",
            "repair_action",
            "original_comment",
            "repaired_comment",
            "repair_status",
            "reason_short",
            "route_to",
            "model_source",
        ])
        for r in results:
            w.writerow([
                r.sample_id,
                r.predicted_label,
                r.detection_confidence,
                r.repair_action,
                r.original_comment,
                r.repaired_comment if r.repaired_comment is not None else "",
                r.repair_status,
                r.reason_short,
                r.route_to,
                r.model_source,
            ])


def write_trace(traces: List[RepairTrace], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for t in traces:
            payload = {
                "sample_id": t.sample_id,
                "language": t.language,
                "code": t.code,
                "original_comment": t.original_comment,
                "predicted_label": t.predicted_label,
                "detection_confidence": t.detection_confidence,
                "first_pass_raw": t.first_pass_raw,
                "second_pass_raw": t.second_pass_raw,
                "repair": asdict(t.repair),
            }
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def run_repair(
    dataset_path: Optional[Path] = None,
    detection_trace_path: Path = Path("outputs") / "detection_trace.jsonl",
    output_dir: Path = Path("outputs"),
    limit: Optional[int] = None,
    fail_fast: bool = True,
    progress_callback: Optional[
        Callable[[str, str, Optional[int], Optional[int], str, Optional[str], Optional[Dict[str, Any]]], None]
    ] = None,
) -> Dict[str, Path]:
    ds_path = resolve_dataset_path(dataset_path)
    if not ds_path.exists():
        raise FileNotFoundError(f"Dataset not found: {ds_path}")
    if not detection_trace_path.exists():
        raise FileNotFoundError(f"Detection trace not found: {detection_trace_path}")

    dataset_map = load_dataset_map(ds_path, limit=limit)
    detection_rows = load_detection_trace(detection_trace_path, limit=limit)
    total = len(detection_rows)

    agent = RepairAgent()
    results: List[RepairOutput] = []
    traces: List[RepairTrace] = []

    if progress_callback is not None:
        progress_callback(
            "repair",
            "running",
            0,
            total,
            "Repair stage started",
            None,
            {"detection_trace": str(detection_trace_path)},
        )

    for idx, rec in enumerate(detection_rows, start=1):
        sid = str(rec.get("sample_id", ""))
        ds_row = dataset_map.get(sid, {"sample_id": sid, "language": "", "code": "", "comment": "", "gold_label": ""})
        try:
            out, first_raw, second_raw = agent.repair_one(rec, ds_row)
        except (LLMRequestError, ValueError, TypeError, KeyError) as exc:
            if progress_callback is not None:
                progress_callback(
                    "repair",
                    "failed",
                    idx,
                    total,
                    "Repair failed on sample",
                    sid,
                    {"error": str(exc)},
                )
            if fail_fast:
                raise RuntimeError(f"Repair failed on sample_id={sid}: {exc}") from exc
            raise

        results.append(out)
        traces.append(
            RepairTrace(
                sample_id=out.sample_id,
                language=str(rec.get("language", ds_row.get("language", ""))),
                code=str(rec.get("code", ds_row.get("code", ""))),
                original_comment=str(rec.get("comment", ds_row.get("comment", ""))),
                predicted_label=out.predicted_label,
                detection_confidence=out.detection_confidence,
                first_pass_raw=first_raw,
                second_pass_raw=second_raw,
                repair=out,
            )
        )
        if progress_callback is not None:
            progress_callback(
                "repair",
                "running",
                idx,
                total,
                "Repair sample processed",
                sid,
                {"repair_action": out.repair_action, "repair_status": out.repair_status},
            )

    jsonl_path = output_dir / "repair_results.jsonl"
    csv_path = output_dir / "repair_results.csv"
    trace_path = output_dir / "repair_trace.jsonl"

    write_outputs(results, jsonl_path, csv_path)
    write_trace(traces, trace_path)
    if progress_callback is not None:
        progress_callback(
            "repair",
            "completed",
            total,
            total,
            "Repair stage completed",
            None,
            {"results_jsonl": str(jsonl_path), "trace_jsonl": str(trace_path)},
        )

    return {
        "repair_results_jsonl": jsonl_path,
        "repair_results_csv": csv_path,
        "repair_trace_jsonl": trace_path,
    }


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Run Repair Agent on detection_trace + Dataset.xlsx")
    p.add_argument("--dataset", type=Path, default=None)
    p.add_argument("--detection-trace", type=Path, default=Path("outputs") / "detection_trace.jsonl")
    p.add_argument("--output-dir", type=Path, default=Path("outputs"))
    p.add_argument("--limit", type=int, default=None)
    return p.parse_args()


def main() -> None:
    args = parse_args()
    run_repair(
        dataset_path=args.dataset,
        detection_trace_path=args.detection_trace,
        output_dir=args.output_dir,
        limit=args.limit,
    )


if __name__ == "__main__":
    main()
