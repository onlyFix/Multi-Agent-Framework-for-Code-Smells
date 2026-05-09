#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from Detection_Agent import ALL_LABELS, DetectionAgent, NOT_SMELL, canon_label
from llm_client import DeepSeekClient, LLMRequestError


@dataclass
class SampleEval:
    sample_id: str
    case_type: str
    failure_stage: str
    repair_attempted: bool
    repair_evaluable: bool
    closed_loop_evaluable: bool
    detection_correctness: bool
    calibration_score: float
    repair_quality_score: Optional[float]
    closed_loop_pass: Optional[bool]
    final_verdict: str
    predicted_label: str
    gold_label: str
    binary_gold_label: str
    binary_predicted_label: str
    binary_detection_correctness: bool
    detection_confidence: float
    repair_action: str
    repair_status: str
    lexical_similarity_score: Optional[float]
    post_repair_label: Optional[str]
    post_repair_confidence: Optional[float]
    smell_removed: Optional[bool]
    new_smell_introduced: Optional[bool]
    label_transition: Optional[str]
    llm_judge_score: Optional[float]
    second_pass_used: bool



def resolve_dataset_path(path: Optional[Path] = None) -> Path:
    if path is not None:
        return path
    p1 = Path("Data") / "Dataset.xlsx"
    p2 = Path("data") / "Dataset.xlsx"
    if p1.exists():
        return p1
    return p2


def norm(s: Optional[str]) -> str:
    return "" if s is None else str(s).strip()


def find_headers(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        mapping[str(v).strip().lower()] = col
    return mapping


def get_cell(ws, row: int, headers: Dict[str, int], names: Sequence[str]) -> str:
    for name in names:
        idx = headers.get(name.lower())
        if idx is None:
            continue
        v = ws.cell(row=row, column=idx).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def load_dataset_map(path: Path, limit: Optional[int] = None) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = find_headers(ws)

    out: Dict[str, Dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        sid = get_cell(ws, row, headers, ["sample_id", "id"])
        if not sid:
            sid = str(row - 1)
        out[sid] = {
            "sample_id": sid,
            "language": get_cell(ws, row, headers, ["language", "lang"]),
            "code": get_cell(ws, row, headers, ["code"]),
            "comment": get_cell(ws, row, headers, ["comment"]),
            "gold_label": get_cell(ws, row, headers, ["gold_label"]),
            "binary_label": get_cell(ws, row, headers, ["binary_label"]),
            "repair_action": get_cell(ws, row, headers, ["repair_action"]),
            "gold_repair": get_cell(ws, row, headers, ["gold_repair"]),
        }
        if limit is not None and len(out) >= limit:
            break
    return out


def load_jsonl_by_id(path: Path, limit: Optional[int] = None) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            sid = str(obj.get("sample_id", ""))
            out[sid] = obj
            if limit is not None and len(out) >= limit:
                break
    return out


def lexical_f1(a: str, b: str) -> float:
    sa = set(norm(a).lower().split())
    sb = set(norm(b).lower().split())
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    inter = len(sa & sb)
    p = inter / len(sa)
    r = inter / len(sb)
    return (2 * p * r / (p + r)) if (p + r) else 0.0


def brier_binary(p: float, y: int) -> float:
    return (p - y) ** 2


def calibration_proxy(confidence: float, correct: bool) -> float:
    return 1.0 - brier_binary(confidence, 1 if correct else 0)


def to_binary_label(label: str) -> str:
    return NOT_SMELL if canon_label(label) == NOT_SMELL else "Smell"


def _round4(x: float) -> float:
    return round(float(x), 4)


def compute_confusion_matrix(gold_labels: List[str], predicted_labels: List[str], label_list: List[str]) -> Dict[str, Dict[str, int]]:
    matrix: Dict[str, Dict[str, int]] = {g: {p: 0 for p in label_list} for g in label_list}
    for g, p in zip(gold_labels, predicted_labels):
        if g not in matrix:
            continue
        if p not in matrix[g]:
            continue
        matrix[g][p] += 1
    return matrix


def compute_per_class_prf(confusion_matrix: Dict[str, Dict[str, int]], label_list: List[str]) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for label in label_list:
        tp = confusion_matrix.get(label, {}).get(label, 0)
        fp = sum(confusion_matrix.get(g, {}).get(label, 0) for g in label_list if g != label)
        fn = sum(confusion_matrix.get(label, {}).get(p, 0) for p in label_list if p != label)
        support = sum(confusion_matrix.get(label, {}).values())

        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

        out[label] = {
            "precision": _round4(precision),
            "recall": _round4(recall),
            "f1": _round4(f1),
            "support": int(support),
        }
    return out


def compute_macro_f1(per_class_prf: Dict[str, Dict[str, float]]) -> float:
    if not per_class_prf:
        return 0.0
    vals = [float(v.get("f1", 0.0)) for v in per_class_prf.values()]
    return _round4(sum(vals) / len(vals)) if vals else 0.0


def compute_balanced_accuracy(per_class_prf: Dict[str, Dict[str, float]]) -> float:
    if not per_class_prf:
        return 0.0
    recalls = [float(v.get("recall", 0.0)) for v in per_class_prf.values()]
    return _round4(sum(recalls) / len(recalls)) if recalls else 0.0


def compute_mcc_multiclass(gold_labels: List[str], predicted_labels: List[str], label_list: List[str]) -> float:
    if not gold_labels or not predicted_labels or not label_list:
        return 0.0

    idx = {label: i for i, label in enumerate(label_list)}
    n = len(label_list)
    cm = [[0 for _ in range(n)] for _ in range(n)]
    for g, p in zip(gold_labels, predicted_labels):
        if g in idx and p in idx:
            cm[idx[g]][idx[p]] += 1

    t_k = [sum(row) for row in cm]
    p_k = [sum(cm[i][j] for i in range(n)) for j in range(n)]
    c = sum(cm[i][i] for i in range(n))
    s = sum(t_k)
    if s == 0:
        return 0.0

    sum_pk_tk = sum(p_k[i] * t_k[i] for i in range(n))
    num = (c * s) - sum_pk_tk
    den_left = (s * s) - sum(v * v for v in p_k)
    den_right = (s * s) - sum(v * v for v in t_k)
    den = math.sqrt(max(den_left, 0.0) * max(den_right, 0.0))
    if den == 0:
        return 0.0
    return _round4(num / den)


def compute_binary_metrics(rows: List[SampleEval]) -> Dict[str, float]:
    rows_with_gold = [r for r in rows if r.binary_gold_label in {"Smell", NOT_SMELL}]
    total = len(rows_with_gold)
    if total == 0:
        return {
            "accuracy": 0.0,
            "precision": 0.0,
            "recall": 0.0,
            "f1": 0.0,
        }

    tp = sum(1 for r in rows_with_gold if r.binary_gold_label == "Smell" and r.binary_predicted_label == "Smell")
    tn = sum(1 for r in rows_with_gold if r.binary_gold_label == NOT_SMELL and r.binary_predicted_label == NOT_SMELL)
    fp = sum(1 for r in rows_with_gold if r.binary_gold_label == NOT_SMELL and r.binary_predicted_label == "Smell")
    fn = sum(1 for r in rows_with_gold if r.binary_gold_label == "Smell" and r.binary_predicted_label == NOT_SMELL)

    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
    acc = (tp + tn) / total if total else 0.0
    return {
        "accuracy": _round4(acc),
        "precision": _round4(precision),
        "recall": _round4(recall),
        "f1": _round4(f1),
    }


def compute_ece(rows: List[SampleEval], n_bins: int = 10) -> float:
    if not rows:
        return 0.0
    bins: List[List[SampleEval]] = [[] for _ in range(n_bins)]
    for r in rows:
        c = max(0.0, min(1.0, float(r.detection_confidence)))
        idx = min(int(c * n_bins), n_bins - 1)
        bins[idx].append(r)

    total = len(rows)
    ece = 0.0
    for b in bins:
        if not b:
            continue
        avg_conf = sum(float(x.detection_confidence) for x in b) / len(b)
        acc = sum(1 for x in b if x.detection_correctness) / len(b)
        ece += (len(b) / total) * abs(acc - avg_conf)
    return _round4(ece)


def average_by(rows: List[SampleEval], key_fn: Callable[[SampleEval], str], value_fn: Callable[[SampleEval], float]) -> Dict[str, float]:
    sums: Dict[str, float] = {}
    counts: Dict[str, int] = {}
    for r in rows:
        k = key_fn(r)
        sums[k] = sums.get(k, 0.0) + float(value_fn(r))
        counts[k] = counts.get(k, 0) + 1
    out: Dict[str, float] = {}
    for k, total in sums.items():
        c = counts.get(k, 0)
        out[k] = _round4(total / c) if c else 0.0
    return dict(sorted(out.items(), key=lambda kv: kv[0]))


def distribution_by(rows: List[SampleEval], key_fn: Callable[[SampleEval], str]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for r in rows:
        k = key_fn(r)
        out[k] = out.get(k, 0) + 1
    return dict(sorted(out.items(), key=lambda kv: kv[0]))


class EvaluationAgent:
    def __init__(
        self,
        confidence_threshold: float = 0.65,
        repair_quality_threshold: float = 0.70,
    ) -> None:
        self.client = DeepSeekClient()
        self.detector = DetectionAgent(confidence_threshold=confidence_threshold)
        self.confidence_threshold = confidence_threshold
        self.repair_quality_threshold = repair_quality_threshold

    def llm_judge(self, code: str, original_comment: str, repaired_comment: str, predicted_label: str) -> Tuple[float, str]:
        system_prompt = "You are an automatic repair quality judge. Return JSON only."
        user_prompt = f"""
Score repaired comment quality from 0 to 1.
Return JSON fields:
- score
- reason_short

Consider:
1) Faithful to code
2) Smell reduction for predicted label
3) Clarity and conciseness

predicted_label: {predicted_label}

code:
{code}

original_comment:
{original_comment}

repaired_comment:
{repaired_comment}
""".strip()
        obj = self.client.chat_json(system_prompt, user_prompt, temperature=0.0, max_tokens=300)
        score = float(obj.get("score", 0.5))
        reason = norm(obj.get("reason_short"))
        return max(0.0, min(1.0, score)), reason

    def second_pass_eval(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        system_prompt = "You are a second-pass evaluator. Return JSON only."
        user_prompt = f"""
Given first-pass evaluation signals, return final verdict.
Allowed final_verdict: ["ACCEPTED", "FAILED"]

Input:
{json.dumps(payload, ensure_ascii=False)}

Return JSON:
{{"final_verdict":"ACCEPTED|FAILED","reason_short":"..."}}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=0.0, max_tokens=250)

    def evaluate_one(self, ds: Dict[str, str], det_record: Dict[str, Any], rep_record: Dict[str, Any]) -> Tuple[SampleEval, Dict[str, Any]]:
        det = det_record.get("detection", {})
        rep = rep_record.get("repair", {})

        sid = ds.get("sample_id", "")
        gold_label = canon_label(ds.get("gold_label", "")) if ds.get("gold_label") else ""
        predicted_label = canon_label(norm(det.get("label", "Not a smell")))
        confidence = float(det.get("confidence", 0.0))
        binary_gold_label = to_binary_label(gold_label) if gold_label else ""
        binary_predicted_label = to_binary_label(predicted_label)

        repair_action = norm(rep.get("repair_action"))
        repair_status = norm(rep.get("repair_status"))
        original_comment = norm(rep.get("original_comment") or ds.get("comment", ""))
        repaired_comment = rep.get("repaired_comment")
        repaired_text = "" if repaired_comment is None else str(repaired_comment)

        detection_correctness = bool(gold_label and predicted_label == gold_label)
        binary_detection_correctness = bool(binary_gold_label and binary_gold_label == binary_predicted_label)
        calibration = calibration_proxy(confidence, detection_correctness)

        gold_is_clean = gold_label == NOT_SMELL
        pred_is_clean = predicted_label == NOT_SMELL

        if gold_is_clean and pred_is_clean:
            case_type = "clean_correct"
        elif gold_is_clean and (not pred_is_clean):
            case_type = "false_positive"
        elif (not gold_is_clean) and pred_is_clean:
            case_type = "false_negative"
        elif (not gold_is_clean) and (not pred_is_clean) and (predicted_label != gold_label):
            case_type = "wrong_smell_type"
        elif (not gold_is_clean) and (predicted_label == gold_label):
            case_type = "repair_eligible"
        else:
            case_type = "other"

        repair_attempted = False
        repair_evaluable = False
        closed_loop_evaluable = False
        repair_quality: Optional[float] = None
        lexical_score: Optional[float] = None
        judge_score: Optional[float] = None
        judge_reason = ""
        post_label: Optional[str] = None
        post_conf: Optional[float] = None
        smell_removed: Optional[bool] = None
        new_smell_introduced: Optional[bool] = None
        label_transition: Optional[str] = None
        closed_loop_pass: Optional[bool] = None
        second_pass_used = False
        failure_stage = ""
        final_verdict = "FAILED"

        if case_type == "clean_correct":
            final_verdict = "ACCEPTED"
        elif case_type in {"false_negative", "false_positive", "wrong_smell_type"}:
            final_verdict = "FAILED"
            failure_stage = "detection"
            repair_attempted = repair_status in {
                "repaired",
                "repaired_second_pass",
                "removed",
                "moved",
                "fallback_original",
            }
        elif case_type == "repair_eligible":
            repair_evaluable = True
            closed_loop_evaluable = True

            # fallback_original means the repair model failed to produce a valid repair.
            # It should not be counted as an effective repair attempt.
            repair_attempted = repair_status in {
                "repaired",
                "repaired_second_pass",
                "removed",
                "moved",
            }

            # Case 1: the repair model failed and fell back to the original comment.
            # Do not judge or re-detect it as if it were a real repair.
            if repair_status == "fallback_original":
                repair_quality = 0.0
                lexical_score = None
                judge_score = None
                judge_reason = "Repair skipped: model returned empty output and fell back to the original comment."
                post_label = predicted_label
                post_conf = confidence
                smell_removed = False
                new_smell_introduced = False
                label_transition = f"{predicted_label} -> {post_label}"
                closed_loop_pass = False
                final_verdict = "FAILED"
                failure_stage = "repair"
                second_pass_used = False

            # Case 2: deletion-like repairs.
            # For Remove/Move, an empty repaired_comment is the expected output.
            # Therefore, do not call llm_judge on the empty string.
            elif repair_action in {"Remove", "Move"}:
                lexical_score = None
                judge_score = None

                if repair_action == "Remove":
                    judge_reason = "LLM judge skipped: Remove action intentionally deletes the inline comment."
                else:
                    judge_reason = "LLM judge skipped: Move action removes the nonlocal note from the inline comment."

                post_label = NOT_SMELL
                post_conf = 1.0
                smell_removed = predicted_label != NOT_SMELL
                new_smell_introduced = False
                label_transition = f"{predicted_label} -> {post_label}"
                closed_loop_pass = True

                repair_quality = 1.0 if (
                    detection_correctness
                    and closed_loop_pass
                    and not new_smell_introduced
                ) else 0.0

                first_pass_accept = bool(
                    detection_correctness
                    and closed_loop_pass
                    and not new_smell_introduced
                )

                final_verdict = "ACCEPTED" if first_pass_accept else "FAILED"

                if not first_pass_accept:
                    failure_stage = "repair_or_closed_loop"

            # Case 3: text-producing repairs, such as Rewrite or Summarize.
            # Only these should be evaluated by LLM judge.
            else:
                gold_repair = norm(ds.get("gold_repair", ""))

                # Lexical similarity is only a diagnostic metric.
                # If no gold repair exists, do not invent a pseudo lexical score.
                if gold_repair:
                    lexical_score = lexical_f1(repaired_text, gold_repair)
                else:
                    lexical_score = None

                if not norm(repaired_text):
                    repair_quality = 0.0
                    judge_score = None
                    judge_reason = "Repair failed: text-producing action returned an empty repaired comment."
                    post_label = predicted_label
                    post_conf = confidence
                    smell_removed = False
                    new_smell_introduced = False
                    label_transition = f"{predicted_label} -> {post_label}"
                    closed_loop_pass = False
                    final_verdict = "FAILED"
                    failure_stage = "repair"
                    second_pass_used = False

                else:
                    try:
                        judge_score, judge_reason = self.llm_judge(
                            ds.get("code", ""),
                            original_comment,
                            repaired_text,
                            predicted_label,
                        )
                    except LLMRequestError as exc:
                        judge_score = 0.0
                        judge_reason = f"judge_failed: {str(exc)[:180]}"

                    # Main repair quality should be based on LLM judge.
                    # lexical_score is reported separately only.
                    repair_quality = max(
                        0.0,
                        min(1.0, judge_score or 0.0),
                    )

                    post_comment = repaired_text
                    post_input = {
                        "sample_id": sid,
                        "language": ds.get("language", ""),
                        "code": ds.get("code", ""),
                        "comment": post_comment,
                    }
                    post_out, _, _, _ = self.detector.detect_one(post_input)
                    post_label = post_out.label
                    post_conf = post_out.confidence
                    smell_removed = predicted_label != NOT_SMELL and post_label == NOT_SMELL
                    new_smell_introduced = (
                        predicted_label != NOT_SMELL
                        and post_label != NOT_SMELL
                        and post_label != predicted_label
                    )
                    label_transition = f"{predicted_label} -> {post_label}"
                    closed_loop_pass = post_label == NOT_SMELL

                    first_pass_accept = (
                        confidence >= self.confidence_threshold
                        and (repair_quality is not None and repair_quality >= self.repair_quality_threshold)
                        and bool(closed_loop_pass)
                    )

                    final_verdict = "ACCEPTED" if first_pass_accept else "FAILED"

                    if not first_pass_accept:
                        failure_stage = "repair_or_closed_loop"
                        second_pass_used = True
                        payload = {
                            "detection_correctness": detection_correctness,
                            "confidence": confidence,
                            "repair_quality_score": round(repair_quality or 0.0, 4),
                            "closed_loop_pass": closed_loop_pass,
                            "predicted_label": predicted_label,
                            "post_repair_label": post_label,
                            "repair_action": repair_action,
                            "repair_status": repair_status,
                        }
                        try:
                            sec = self.second_pass_eval(payload)
                            sec_verdict = norm(sec.get("final_verdict", "FAILED")).upper()
                            if sec_verdict in {"ACCEPTED", "FAILED"}:
                                final_verdict = sec_verdict
                            judge_reason = (
                                judge_reason
                                + " | second_pass: "
                                + norm(sec.get("reason_short"))
                            ).strip(" |")
                        except LLMRequestError:
                            pass
        else:
            final_verdict = "FAILED"
            failure_stage = "unknown"

        if final_verdict == "FAILED" and not failure_stage:
            failure_stage = "detection"

        row = SampleEval(
            sample_id=sid,
            case_type=case_type,
            failure_stage=failure_stage,
            repair_attempted=repair_attempted,
            repair_evaluable=repair_evaluable,
            closed_loop_evaluable=closed_loop_evaluable,
            detection_correctness=detection_correctness,
            calibration_score=round(calibration, 4),
            repair_quality_score=(round(repair_quality, 4) if repair_quality is not None else None),
            closed_loop_pass=closed_loop_pass,
            final_verdict=final_verdict,
            predicted_label=predicted_label,
            gold_label=gold_label,
            binary_gold_label=binary_gold_label,
            binary_predicted_label=binary_predicted_label,
            binary_detection_correctness=binary_detection_correctness,
            detection_confidence=round(confidence, 4),
            repair_action=repair_action,
            repair_status=repair_status,
            lexical_similarity_score=(round(lexical_score, 4) if lexical_score is not None else None),
            post_repair_label=post_label,
            post_repair_confidence=(round(post_conf, 4) if post_conf is not None else None),
            smell_removed=smell_removed,
            new_smell_introduced=new_smell_introduced,
            label_transition=label_transition,
            llm_judge_score=(round(judge_score, 4) if judge_score is not None else None),
            second_pass_used=second_pass_used,
        )

        extra = {
            "sample_id": sid,
            "judge_reason_short": judge_reason,
        }
        return row, extra


def write_per_sample(rows: List[SampleEval], jsonl_path: Path, csv_path: Path) -> None:
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    with jsonl_path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(asdict(r), ensure_ascii=False) + "\n")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "sample_id",
            "case_type",
            "failure_stage",
            "repair_attempted",
            "repair_evaluable",
            "closed_loop_evaluable",
            "detection_correctness",
            "calibration_score",
            "repair_quality_score",
            "closed_loop_pass",
            "final_verdict",
            "predicted_label",
            "gold_label",
            "binary_gold_label",
            "binary_predicted_label",
            "binary_detection_correctness",
            "detection_confidence",
            "repair_action",
            "repair_status",
            "lexical_similarity_score",
            "post_repair_label",
            "post_repair_confidence",
            "smell_removed",
            "new_smell_introduced",
            "label_transition",
            "llm_judge_score",
            "second_pass_used",
        ])
        for r in rows:
            w.writerow([
                r.sample_id,
                r.case_type,
                r.failure_stage,
                r.repair_attempted,
                r.repair_evaluable,
                r.closed_loop_evaluable,
                r.detection_correctness,
                r.calibration_score,
                r.repair_quality_score,
                r.closed_loop_pass,
                r.final_verdict,
                r.predicted_label,
                r.gold_label,
                r.binary_gold_label,
                r.binary_predicted_label,
                r.binary_detection_correctness,
                r.detection_confidence,
                r.repair_action,
                r.repair_status,
                r.lexical_similarity_score,
                r.post_repair_label,
                r.post_repair_confidence,
                r.smell_removed,
                r.new_smell_introduced,
                r.label_transition,
                r.llm_judge_score,
                r.second_pass_used,
            ])


def write_report(report: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def run_evaluation(
    dataset_path: Optional[Path] = None,
    detection_trace_path: Path = Path("outputs") / "detection_trace.jsonl",
    repair_trace_path: Path = Path("outputs") / "repair_trace.jsonl",
    output_dir: Path = Path("outputs"),
    limit: Optional[int] = None,
    confidence_threshold: float = 0.65,
    repair_quality_threshold: float = 0.70,
    progress_callback: Optional[
        Callable[[str, str, Optional[int], Optional[int], str, Optional[str], Optional[Dict[str, Any]]], None]
    ] = None,
) -> Dict[str, Path]:
    ds_path = resolve_dataset_path(dataset_path)
    if not ds_path.exists():
        raise FileNotFoundError(f"Dataset not found: {ds_path}")
    if not detection_trace_path.exists():
        raise FileNotFoundError(f"Detection trace not found: {detection_trace_path}")
    if not repair_trace_path.exists():
        raise FileNotFoundError(f"Repair trace not found: {repair_trace_path}")

    dataset = load_dataset_map(ds_path, limit=limit)
    det_map = load_jsonl_by_id(detection_trace_path, limit=limit)
    rep_map = load_jsonl_by_id(repair_trace_path, limit=limit)

    shared_ids = [sid for sid in dataset.keys() if sid in det_map and sid in rep_map]
    total = len(shared_ids)

    agent = EvaluationAgent(
        confidence_threshold=confidence_threshold,
        repair_quality_threshold=repair_quality_threshold,
    )
    per_sample: List[SampleEval] = []
    extras: List[Dict[str, Any]] = []

    if progress_callback is not None:
        progress_callback(
            "evaluation",
            "running",
            0,
            total,
            "Evaluation stage started",
            None,
            {"repair_trace": str(repair_trace_path)},
        )

    for idx, sid in enumerate(shared_ids, start=1):
        row, extra = agent.evaluate_one(dataset[sid], det_map[sid], rep_map[sid])
        per_sample.append(row)
        extras.append(extra)
        if progress_callback is not None:
            progress_callback(
                "evaluation",
                "running",
                idx,
                total,
                "Evaluation sample processed",
                sid,
                {"final_verdict": row.final_verdict},
            )

    total = len(per_sample)
    accepted = sum(1 for r in per_sample if r.final_verdict == "ACCEPTED")
    failed = total - accepted

    eval_rows_with_gold = [r for r in per_sample if r.gold_label]
    gold_labels = [r.gold_label for r in eval_rows_with_gold]
    predicted_labels = [r.predicted_label for r in eval_rows_with_gold]

    confusion = compute_confusion_matrix(gold_labels, predicted_labels, ALL_LABELS)
    per_class_prf = compute_per_class_prf(confusion, ALL_LABELS)
    macro_f1 = compute_macro_f1(per_class_prf)
    balanced_acc = compute_balanced_accuracy(per_class_prf)
    mcc = compute_mcc_multiclass(gold_labels, predicted_labels, ALL_LABELS)

    det_acc = (sum(1 for r in eval_rows_with_gold if r.detection_correctness) / len(eval_rows_with_gold)) if eval_rows_with_gold else 0.0
    binary_metrics = compute_binary_metrics(eval_rows_with_gold)

    avg_calib = sum(r.calibration_score for r in per_sample) / total if total else 0.0
    rows_correct = [r for r in per_sample if r.detection_correctness]
    rows_wrong = [r for r in per_sample if not r.detection_correctness]
    avg_conf_correct = (sum(r.detection_confidence for r in rows_correct) / len(rows_correct)) if rows_correct else 0.0
    avg_conf_wrong = (sum(r.detection_confidence for r in rows_wrong) / len(rows_wrong)) if rows_wrong else 0.0
    second_check_triggered = sum(
        1
        for sid in shared_ids
        if det_map.get(sid, {}).get("second_check_raw") is not None
    )
    second_check_trigger_rate = (second_check_triggered / total) if total else 0.0
    ece = compute_ece(per_sample, n_bins=10)

    repair_rows = [r for r in per_sample if r.repair_evaluable and r.repair_quality_score is not None]
    closed_loop_rows = [r for r in per_sample if r.closed_loop_evaluable and r.closed_loop_pass is not None]
    reference_rows = [
        r for r in repair_rows
        if r.lexical_similarity_score is not None
    ]

    avg_repair = sum(float(r.repair_quality_score or 0.0) for r in repair_rows) / len(repair_rows) if repair_rows else 0.0
    avg_llm_judge = sum(float(r.llm_judge_score or 0.0) for r in repair_rows) / len(repair_rows) if repair_rows else 0.0
    avg_ref_sim = (
        sum(float(r.lexical_similarity_score) for r in reference_rows) / len(reference_rows)
        if reference_rows else 0.0
    )
    repair_by_label = average_by(repair_rows, lambda r: r.predicted_label, lambda r: float(r.repair_quality_score or 0.0))
    llm_judge_by_label = average_by(repair_rows, lambda r: r.predicted_label, lambda r: float(r.llm_judge_score or 0.0))
    lexical_by_label = average_by(
        reference_rows,
        lambda r: r.predicted_label,
        lambda r: float(r.lexical_similarity_score or 0.0),
    )
    repair_by_action = average_by(repair_rows, lambda r: r.repair_action, lambda r: float(r.repair_quality_score or 0.0))
    repair_status_dist = distribution_by(repair_rows, lambda r: r.repair_status)

    closed_loop_rate = (
        sum(1 for r in closed_loop_rows if r.closed_loop_pass) / len(closed_loop_rows)
    ) if closed_loop_rows else 0.0
    smell_removed_rate = (
        sum(1 for r in closed_loop_rows if r.smell_removed) / len(closed_loop_rows)
    ) if closed_loop_rows else 0.0
    residual_smell_rate = (
        sum(1 for r in closed_loop_rows if r.predicted_label != NOT_SMELL and r.post_repair_label != NOT_SMELL) / len(closed_loop_rows)
    ) if closed_loop_rows else 0.0
    new_smell_rate = (
        sum(1 for r in closed_loop_rows if r.new_smell_introduced) / len(closed_loop_rows)
    ) if closed_loop_rows else 0.0
    pass_rate_by_label = average_by(
        closed_loop_rows,
        lambda r: r.predicted_label,
        lambda r: 1.0 if r.closed_loop_pass else 0.0,  # type: ignore[arg-type]
    )
    label_transition_dist = distribution_by(
        [r for r in closed_loop_rows if r.label_transition],
        lambda r: str(r.label_transition),
    )
    second_pass_usage_rate = (sum(1 for r in repair_rows if r.second_pass_used) / len(repair_rows)) if repair_rows else 0.0

    report = {
        "metadata": {
            "dataset": str(ds_path),
            "detection_trace": str(detection_trace_path),
            "repair_trace": str(repair_trace_path),
            "sample_count": total,
            "model_source": agent.client.model_source,
        },
        "detection_metrics": {
            "accuracy": _round4(det_acc),
            "macro_f1": macro_f1,
            "mcc": mcc,
            "balanced_accuracy": balanced_acc,
            "binary_accuracy": binary_metrics["accuracy"],
            "binary_precision": binary_metrics["precision"],
            "binary_recall": binary_metrics["recall"],
            "binary_f1": binary_metrics["f1"],
            "per_class_prf": per_class_prf,
            "confusion_matrix": confusion,
        },
        "calibration_metrics": {
            "avg_calibration_score": _round4(avg_calib),
            "avg_confidence_correct": _round4(avg_conf_correct),
            "avg_confidence_wrong": _round4(avg_conf_wrong),
            "second_check_trigger_rate": _round4(second_check_trigger_rate),
            "ece": ece,
        },
        "repair_metrics": {
            "repair_evaluable_count": len(repair_rows),
            "avg_repair_quality_score": _round4(avg_repair),
            "avg_llm_judge_score": _round4(avg_llm_judge),
            "avg_reference_similarity": _round4(avg_ref_sim),
            "avg_repair_quality_by_label": repair_by_label,
            "avg_llm_judge_score_by_label": llm_judge_by_label,
            "avg_reference_similarity_by_label": lexical_by_label,
            "avg_repair_quality_by_action": repair_by_action,
            "repair_status_distribution": repair_status_dist,
        },
        "closed_loop_metrics": {
            "closed_loop_evaluable_count": len(closed_loop_rows),
            "closed_loop_pass_rate": _round4(closed_loop_rate),
            "smell_removed_rate": _round4(smell_removed_rate),
            "residual_smell_rate": _round4(residual_smell_rate),
            "new_smell_introduction_rate": _round4(new_smell_rate),
            "closed_loop_pass_rate_by_label": pass_rate_by_label,
            "label_transition_distribution": label_transition_dist,
            "second_pass_usage_rate": _round4(second_pass_usage_rate),
        },
        "final_verdict_distribution": {
            "ACCEPTED": accepted,
            "FAILED": failed,
        },
        "notes": extras,
    }

    report_path = output_dir / "evaluation_report.json"
    per_jsonl = output_dir / "evaluation_per_sample.jsonl"
    per_csv = output_dir / "evaluation_per_sample.csv"

    write_report(report, report_path)
    write_per_sample(per_sample, per_jsonl, per_csv)
    if progress_callback is not None:
        progress_callback(
            "evaluation",
            "completed",
            total,
            total,
            "Evaluation stage completed",
            None,
            {"report_json": str(report_path)},
        )

    return {
        "evaluation_report_json": report_path,
        "evaluation_per_sample_jsonl": per_jsonl,
        "evaluation_per_sample_csv": per_csv,
    }


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Run Evaluation Agent with automatic second-pass decisions")
    p.add_argument("--dataset", type=Path, default=None)
    p.add_argument("--detection-trace", type=Path, default=Path("outputs") / "detection_trace.jsonl")
    p.add_argument("--repair-trace", type=Path, default=Path("outputs") / "repair_trace.jsonl")
    p.add_argument("--output-dir", type=Path, default=Path("outputs"))
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--confidence-threshold", type=float, default=0.65)
    p.add_argument("--repair-quality-threshold", type=float, default=0.70)
    return p.parse_args()


def main() -> None:
    args = parse_args()
    run_evaluation(
        dataset_path=args.dataset,
        detection_trace_path=args.detection_trace,
        repair_trace_path=args.repair_trace,
        output_dir=args.output_dir,
        limit=args.limit,
        confidence_threshold=args.confidence_threshold,
        repair_quality_threshold=args.repair_quality_threshold,
    )


if __name__ == "__main__":
    main()
