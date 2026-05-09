#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass, fields
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from llm_client import DeepSeekClient, LLMConfigError, LLMRequestError


LABELS_SMELL = [
    "Misleading",
    "Obvious",
    "Vague",
    "Too much info",
    "Task",
    "Commented out code",
    "Irrelevant",
    "Nonlocal info",
    "Beautification",
    "Attribution",
]
NOT_SMELL = "Not a smell"
ALL_LABELS = LABELS_SMELL + [NOT_SMELL]

TAXONOMY: Dict[str, str] = {
    "Misleading": "Comment says something that's inconsistent or incorrect relative to what the code does.",
    "Obvious": "Comments that merely restate the code without adding useful information.",
    "Commented out code": "A code piece that is commented out.",
    "Irrelevant": "Comments that do not intend to explain the code.",
    "Task": "Comments explaining the work that could/should be done in future or was already completed.",
    "Too much info": "Comments that are overly detailed or verbose, usually stating implementation-level trivia.",
    "Beautification": "Comments that aim to distinguish the parts of the code.",
    "Nonlocal info": "Comments that provide systemwide information or mention code that is not near.",
    "Vague": "Comments that are not clearly understandable.",
    "Attribution": "Comments that give information about who wrote the code.",
    "Not a smell": "Comments that enhance code comprehension, provide context or explain complex logic.",
}

ACTION_HINTS = {
    "Misleading": "rewrite",
    "Obvious": "remove",
    "Vague": "rewrite",
    "Too much info": "summarize",
    "Task": "remove",
    "Commented out code": "remove",
    "Irrelevant": "remove",
    "Nonlocal info": "move",
    "Beautification": "remove",
    "Attribution": "remove",
    "Not a smell": "no_change",
}

EXEMPLAR_POOL: List[Dict[str, str]] = [
    {"comment": "TODO: fix bug when input is null", "code": "if (input == null) { ... }", "label": "Task"},
    {"comment": "// FIXME: this always returns null for empty input", "code": "String getFirstItem(List<String> items) {\n    return items.get(0);\n}", "label": "Task"},
    {"comment": "assigning count a value of 0", "code": "count = 0;", "label": "Obvious"},
    {"comment": "# result = do_something_complex(x, y)", "code": "", "label": "Commented out code"},
    {"comment": "NETWORK CONFIGURATION", "code": "config.load(\"net.cfg\");", "label": "Beautification"},
    {"comment": "Warn if temperature exceeds safe range", "code": "", "label": "Not a smell"},
    {"comment": "Validate user input to ensure non-empty string", "code": "", "label": "Not a smell"},
    {"comment": "Configure the cache size to optimize memory usage", "code": "", "label": "Not a smell"},
    {"comment": "updates user1's properties", "code": "updateDatabase(user2);", "label": "Misleading"},
    {"comment": "Encrypt the password before saving", "code": "user.setPassword(rawPassword);", "label": "Misleading"},
    {"comment": "Fetch user from DB", "code": "User u = userRepository.getById(id);", "label": "Obvious"},
    {"comment": "// if (user.isAdmin() && user.hasAccess()) { grantPermissions(); }", "code": "", "label": "Commented out code"},
    {"comment": "I dedicate this project to my cat, Fluffy", "code": "String name = \"Fluffy\";", "label": "Irrelevant"},
    {"comment": "this method is so bad lol", "code": "handleLogin();", "label": "Irrelevant"},
    {"comment": "This opens the database connection so we can send queries and get results back, assuming credentials are correct", "code": "connection.open();", "label": "Too much info"},
    {"comment": "# Create a list to store user names\n# We'll use a Python list which supports dynamic resizing\n# and provides O(1) append operations under the hood", "code": "user_names = []", "label": "Too much info"},
    {"comment": "// =========================\n// === INITIALIZATION ===\n// =========================", "code": "initializeApp();", "label": "Beautification"},
    {"comment": "Keep timeout low to fail fast on flaky network links", "code": "client.setTimeout(1000);", "label": "Not a smell"},
    {"comment": "Parse server response JSON into model fields", "code": "parseResponse(body);", "label": "Not a smell"},
    {"comment": "This constant is also used in the logging module", "code": "const int MAX_RETRIES = 5;", "label": "Nonlocal info"},
    {"comment": "# This timeout affects the global Redis configuration in config/database.yml", "code": "def set_cache_timeout(seconds):\n    self.timeout = seconds", "label": "Nonlocal info"},
    {"comment": "Do the thing", "code": "result = reconcileUserBalances(accounts);", "label": "Vague"},
    {"comment": "taskTmpPath = ttp; //_task_tmp", "code": "taskTmpPath = ttp;", "label": "Vague"},
    {"comment": "Initialize adapter and bind it to list view", "code": "listView.setAdapter(adapter);", "label": "Obvious"},
    {"comment": "Use stable IDs so item animations stay consistent after reordering", "code": "adapter.setHasStableIds(true);", "label": "Not a smell"},
    {"comment": "Load user profile from local cache to avoid an extra network call", "code": "profile = cache.get(userId);", "label": "Not a smell"},
    {"comment": "This constant is shared across multiple modules and must match server config", "code": "public static final int PORT = 8082;", "label": "Nonlocal info"},
    {"comment": "// Parse server response JSON into model fields", "code": "parseResponse(body);", "label": "Not a smell"},
    {"comment": "// set x to 0", "code": "x = 0;", "label": "Obvious"},
    {"comment": "// Keep this timeout low to fail fast on flaky network links", "code": "client.setTimeout(1000);", "label": "Not a smell"},
    {"comment": "Retry once because this API often clears 429s on the next window", "code": "max_retries = 1", "label": "Not a smell"},
    {"comment": "retry once", "code": "max_retries = 1", "label": "Obvious"},
    {"comment": "Create the cache dir up front because workers assume the path already exists before forking", "code": "cache_dir.mkdir(parents=True, exist_ok=True)", "label": "Not a smell"},
    {"comment": "create the cache directory", "code": "cache_dir.mkdir(parents=True, exist_ok=True)", "label": "Obvious"},
    {"comment": "Treat missing token as anonymous traffic so auth noise is not reported as an error", "code": "if (token == null) return Optional.empty();", "label": "Not a smell"},
    {"comment": "return empty if token is null", "code": "if (token == null) return Optional.empty();", "label": "Obvious"},
]

CONFUSION_PAIRS: List[Tuple[str, str]] = [
    (NOT_SMELL, "Obvious"),
    (NOT_SMELL, "Vague"),
    (NOT_SMELL, "Nonlocal info"),
    (NOT_SMELL, "Too much info"),
    (NOT_SMELL, "Misleading"),
]


@dataclass
class DetectionOutput:
    sample_id: str
    is_smelly: bool
    label: str
    confidence: float
    top2_labels: List[Dict[str, float]]
    reason_short: str
    action_hint: str
    route_to: str
    model_source: str

    
    second_check_used: bool = False
    review_reason: str = ""


@dataclass
class DetectionTrace:
    sample_id: str
    language: str
    code: str
    comment: str
    gold_label: Optional[str]
    stage1_raw: Dict[str, Any]
    stage2_raw: Optional[Dict[str, Any]]
    second_check_raw: Optional[Dict[str, Any]]
    detection: DetectionOutput



def resolve_dataset_path(path: Optional[Path] = None) -> Path:
    if path is not None:
        return path
    p1 = Path("Data") / "Dataset.xlsx"
    p2 = Path("data") / "Dataset.xlsx"
    if p1.exists():
        return p1
    return p2


def norm(s: Optional[str]) -> str:
    return "" if s is None else str(s).strip()


def _tokenize(text: str) -> List[str]:
    return re.findall(r"[a-zA-Z_]{2,}", text.lower())


def _jaccard(a: Sequence[str], b: Sequence[str]) -> float:
    sa, sb = set(a), set(b)
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def _example_similarity(sample_comment: str, sample_code: str, ex_comment: str, ex_code: str) -> float:
    c_sim = _jaccard(_tokenize(sample_comment), _tokenize(ex_comment))
    code_sim = _jaccard(_tokenize(sample_code), _tokenize(ex_code))
    return 0.7 * c_sim + 0.3 * code_sim


def _build_stage1_dynamic_examples(sample: Dict[str, str], k: int = 6) -> str:
    comment = norm(sample.get("comment", ""))
    code = norm(sample.get("code", ""))
    scored: List[Tuple[float, Dict[str, str]]] = []
    for ex in EXEMPLAR_POOL:
        score = _example_similarity(comment, code, ex["comment"], ex["code"])
        scored.append((score, ex))
    scored.sort(key=lambda x: x[0], reverse=True)

    smell_examples = [ex for _, ex in scored if ex["label"] != NOT_SMELL][: max(2, k // 2)]
    clean_examples = [ex for _, ex in scored if ex["label"] == NOT_SMELL][: max(2, k // 2)]
    picked = (smell_examples + clean_examples)[:k]

    lines: List[str] = []
    for idx, ex in enumerate(picked, start=1):
        out_label = "Smell" if ex["label"] != NOT_SMELL else "Not a smell"
        alt_label = "Not a smell" if out_label == "Smell" else "Smell"
        lines.append(
            f'Example {idx}:\n'
            f'Input comment: "{ex["comment"]}"\n'
            f'Output: {{"label":"{out_label}","confidence":0.85,"top2_labels":[{{"label":"{out_label}","score":0.85}},{{"label":"{alt_label}","score":0.15}}],"reason_short":"Reference pattern match."}}'
        )
    return "\n\n".join(lines)


def _build_stage2_dynamic_examples(sample: Dict[str, str], k: int = 8) -> str:
    comment = norm(sample.get("comment", ""))
    code = norm(sample.get("code", ""))
    scored: List[Tuple[float, Dict[str, str]]] = []
    for ex in EXEMPLAR_POOL:
        score = _example_similarity(comment, code, ex["comment"], ex["code"])
        scored.append((score, ex))
    scored.sort(key=lambda x: x[0], reverse=True)

    comment_l = comment.lower()
    comment_words = comment_l.split()

    vague_markers = [
        "this", "it", "thing", "stuff", "process", "handle", "record", "path", "etc"
    ]

    is_short_fragment = len(comment_words) <= 4
    has_vague_marker = any(w in comment_l for w in vague_markers)

    if any(w in comment_l for w in ["module", "config", "deploy", "gateway", "shared", "global", "service", "ops"]):
        focus_pair = (NOT_SMELL, "Nonlocal info")
    elif len(comment.split()) >= 18 or comment.count(";") >= 2:
        focus_pair = (NOT_SMELL, "Too much info")
    elif any(w in comment_l for w in ["descending", "ascending", "invalidate", "append", "before serving"]):
        focus_pair = (NOT_SMELL, "Misleading")
    elif is_short_fragment or has_vague_marker:
        focus_pair = (NOT_SMELL, "Vague")
    else:
        focus_pair = (NOT_SMELL, "Obvious")

    picked: List[Dict[str, str]] = []
    used: set[int] = set()

    def add_best_for_label(label: str, take: int = 1) -> None:
        taken = 0
        for idx, (_, ex) in enumerate(scored):
            if idx in used:
                continue
            if ex["label"] != label:
                continue
            picked.append(ex)
            used.add(idx)
            taken += 1
            if taken >= take:
                break

    add_best_for_label(focus_pair[0], take=2)
    add_best_for_label(focus_pair[1], take=2)
    for a, b in CONFUSION_PAIRS:
        if (a, b) == focus_pair:
            continue
        add_best_for_label(b, take=1)
        if len(picked) >= k - 1:
            break

    for idx, (_, ex) in enumerate(scored):
        if len(picked) >= k:
            break
        if idx in used:
            continue
        picked.append(ex)
        used.add(idx)

    lines: List[str] = []
    for idx, ex in enumerate(picked, start=1):
        lines.append(
            f"Example {idx}:\n"
            f"Code comment: {ex['comment']}\n"
            f"Code segment: {ex['code']}\n"
            f"Label: {ex['label']}"
        )
    return "\n\n".join(lines)


def canon_label(label: str) -> str:
    raw = norm(label)
    low = raw.lower().replace("-", " ")
    aliases = {
        "not a smell": "Not a smell",
        "too much information": "Too much info",
        "too much info": "Too much info",
        "commented out code": "Commented out code",
        "commented-out code": "Commented out code",
        "non local": "Nonlocal info",
        "non-local": "Nonlocal info",
        "nonlocal": "Nonlocal info",
        "nonlocal info": "Nonlocal info",
        "attribution": "Attribution",
    }
    if low in aliases:
        return aliases[low]
    for v in ALL_LABELS:
        if v.lower() == raw.lower():
            return v
    return raw


def find_headers(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        key = str(v).strip().lower()
        mapping[key] = col
    return mapping


def get_cell(ws, row: int, headers: Dict[str, int], names: Sequence[str]) -> str:
    for name in names:
        idx = headers.get(name.lower())
        if idx is None:
            continue
        v = ws.cell(row=row, column=idx).value
        if v is None:
            continue
        text = str(v).strip()
        if text:
            return text
    return ""


def load_samples_from_xlsx(path: Path, limit: Optional[int] = None) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = find_headers(ws)

    out: List[Dict[str, str]] = []
    for row in range(2, ws.max_row + 1):
        sample_id = get_cell(ws, row, headers, ["sample_id", "id"])
        if not sample_id:
            sample_id = str(row - 1)

        rec = {
            "sample_id": sample_id,
            "language": get_cell(ws, row, headers, ["language", "lang"]),
            "code": get_cell(ws, row, headers, ["code"]),
            "comment": get_cell(ws, row, headers, ["comment"]),
            "gold_label": get_cell(ws, row, headers, ["gold_label"]),
            "binary_label": get_cell(ws, row, headers, ["binary_label"]),
            "repair_action": get_cell(ws, row, headers, ["repair_action"]),
            "gold_repair": get_cell(ws, row, headers, ["gold_repair"]),
        }
        out.append(rec)
        if limit is not None and len(out) >= limit:
            break

    return out


class DetectionAgent:
    def __init__(
        self,
        confidence_threshold: float = 0.75,
        second_check_threshold: float = 0.80,
        temperature: float = 0.0,
        borderline_threshold: float = 0.90,
        uncertainty_margin_threshold: float = 0.20,
        hybrid_verify: bool = False,
    ) -> None:
        self.client = DeepSeekClient()
        self.confidence_threshold = confidence_threshold
        self.second_check_threshold = second_check_threshold
        self.temperature = temperature
        self.borderline_threshold = borderline_threshold
        self.uncertainty_margin_threshold = uncertainty_margin_threshold
        self.hybrid_verify = hybrid_verify

    def _extract_top2_labels(self, raw_top2: Any, label: str, conf: float) -> Tuple[str, str]:
        top2 = self._normalize_top2(raw_top2, label, conf)
        first = top2[0]["label"] if top2 else label
        second = top2[1]["label"] if len(top2) > 1 else (NOT_SMELL if first != NOT_SMELL else "Obvious")
        return first, second

    def _signal_conflict(self, label: str, signals: Dict[str, Any]) -> bool:
        has_intent = bool(signals.get("has_rationale_or_intent", False))
        restates = bool(signals.get("mostly_restates_local_code", False))
        ext_ref = bool(signals.get("references_external_module_or_config", False))
        too_detail = bool(signals.get("contains_excess_implementation_detail", False))
        beautify = bool(signals.get("is_section_header_or_visual_separator", False))
        missing_ref = bool(signals.get("missing_specific_referent_or_effect", False))

        if label == "Obvious" and has_intent and not restates:
            return True
        if label == NOT_SMELL and ext_ref:
            return True
        if label == NOT_SMELL and too_detail and not has_intent:
            return True
        if label == NOT_SMELL and beautify:
            return True
        if label == NOT_SMELL and missing_ref:
            return True
        return False

    def _stage1(self, sample: Dict[str, str]) -> Dict[str, Any]:
        dynamic_examples = _build_stage1_dynamic_examples(sample)
        system_prompt = (
            "You are a strict code-comment smell detector. "
            "Classify only with the allowed labels and return JSON only."
        )
        user_prompt = f"""
Task: Stage-1 binary detection.
Allowed labels: ["Smell", "Not a smell"].
Return JSON object with fields:
- label ("Smell" or "Not a smell")
- confidence (0 to 1)
- top2_labels (array of 2 objects: {{"label": str, "score": float}})
- reason_short (<= 25 words)

Taxonomy reminder:
{json.dumps(TAXONOMY, ensure_ascii=False)}

Examples:
{dynamic_examples}

Binary mapping rule:
- If example/analysis label is any smell category (not "Not a smell"), output "Smell".
- Only output "Not a smell" when the comment is genuinely helpful.
- Never output fine-grained labels (e.g., Obvious, Vague) in Stage-1.

Now classify this sample:
language: {sample.get('language', '')}
code:
{sample.get('code', '')}
comment:
{sample.get('comment', '')}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=self.temperature, max_tokens=500)

    def _stage2(self, sample: Dict[str, str]) -> Dict[str, Any]:
        dynamic_examples = _build_stage2_dynamic_examples(sample)
        system_prompt = (
            "You are a strict multi-class code-comment smell classifier. "
            "Return JSON only and labels must be from allowed set."
        )
        user_prompt = f"""
Code comments are considered smells if they degrade software quality or do not help readers with code comprehension.
You will be provided with a taxonomy and examples.

Task: Stage-2 fine-grained smell classification.
Allowed labels: {json.dumps(ALL_LABELS, ensure_ascii=False)}
Return JSON object with fields:
- label (must be one allowed label)
- confidence (0 to 1)
- top2_labels (array of 2 objects: {{"label": str, "score": float}})
- reason_short (<= 25 words)
- signals (object with booleans):
  - has_rationale_or_intent
  - mostly_restates_local_code
  - references_external_module_or_config
  - contains_excess_implementation_detail
  - conflicts_with_code_semantics
  - is_section_header_or_visual_separator
  - missing_specific_referent_or_effect

Taxonomy:
{json.dumps(TAXONOMY, ensure_ascii=False)}

Additional decision policy:
- A short comment is NOT automatically a smell.
- A concise comment can be "Not a smell" when it adds specific local purpose, rationale, constraint, or risk.
- Choose "Obvious" only if comment mainly restates visible local code action with no added why/intent/constraint/risk.
- Choose "Vague" when the comment is too generic, abbreviated, fragmentary, or only names a broad topic without explaining what specifically matters.
- Choose "Vague" if the reader must guess the specific object, action, condition, effect, or purpose.
- Do not upgrade a vague comment to "Not a smell" just because the surrounding code can be inspected. The comment itself must add specific local meaning.
- Choose "Vague" for comments like "handle result", "update state", "check item", "process file", "fix name", or "special handling" when they do not explain the concrete reason or effect.
- Choose "Beautification" only for pure headers/separators with near-zero behavioral/semantic content.
- A comment can be "Not a smell" if it clarifies intent, context, purpose, risk, edge case, or complex logic.
- Prefer "Not a smell" only when the comment adds local purpose, rationale, constraint, or risk and no strong smell evidence is present.
- Do not default to "Not a smell" for TODO/FIXME, commented-out code, contradiction with code, explicit external/system-wide dependency, or clearly underspecified comments.
- If a comment only restates local code without useful context, classify as "Obvious" (not "Not a smell").

Examples:
{dynamic_examples}

Now classify:
language: {sample.get('language', '')}
code:
{sample.get('code', '')}
comment:
{sample.get('comment', '')}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=self.temperature, max_tokens=600)

    def _pair_check(
        self,
        sample: Dict[str, str],
        stage1: Dict[str, Any],
        stage2: Optional[Dict[str, Any]],
        label_a: str,
        label_b: str,
        guidance: str,
    ) -> Dict[str, Any]:
        system_prompt = "You are a strict pairwise tie-breaker for comment classification. Return JSON only."
        user_prompt = f"""
Resolve this confusion pair only: ["{label_a}", "{label_b}"].
Return JSON with fields:
- final_label (must be "{label_a}" or "{label_b}")
- final_confidence (0..1)
- top2_labels (2 items)
- reason_short

Pairwise rules:
{guidance}

Global tie-break policy:
- Do not choose a smell label just because the comment is short.
- If one option is "Not a smell" and comment still provides useful specific context, prefer "Not a smell".
- Choose "Obvious" only for pure local restatement without useful why/intent/constraint/risk.
- Choose "Vague" only when key referent/object/condition/effect remains unclear.

Sample:
language: {sample.get('language', '')}
code:
{sample.get('code', '')}
comment:
{sample.get('comment', '')}

Previous stage1: {json.dumps(stage1, ensure_ascii=False)}
Previous stage2: {json.dumps(stage2, ensure_ascii=False) if stage2 else "null"}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=0.0, max_tokens=350)

    def _pair_check_not_smell_vs_obvious(self, sample: Dict[str, str], stage1: Dict[str, Any], stage2: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        guidance = (
            '- Choose Obvious only if the comment mainly restates the visible local code action and removing it would not lose why, intent, constraint, or risk.\n'
            '- If the comment gives even short but useful rationale, purpose, constraint, or risk for the current code, prefer Not a smell.\n'
            '- Do not choose Obvious just because the comment is short.'
        )
        return self._pair_check(sample, stage1, stage2, NOT_SMELL, "Obvious", guidance)

    def _pair_check_not_smell_vs_vague(
        self,
        sample: Dict[str, str],
        stage1: Dict[str, Any],
        stage2: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        guidance = (
            '- Choose "Vague" if the comment is too generic, abbreviated, fragmentary, or only names a broad topic.\n'
            '- Choose "Vague" if the comment does not clearly identify the specific object, action, condition, effect, or purpose.\n'
            '- Choose "Vague" if the reader must guess what words like "this", "it", "thing", "stuff", "process", "handle", "record", "path", or "etc" refer to.\n'
            '- Choose "Not a smell" only if the comment provides specific local intent, rationale, constraint, risk, edge case, or effect.\n'
            '- Do not choose "Not a smell" just because the code nearby helps infer the meaning; the comment itself must be informative.'
        )
        return self._pair_check(sample, stage1, stage2, NOT_SMELL, "Vague", guidance)

    def _pair_check_not_smell_vs_nonlocal(self, sample: Dict[str, str], stage1: Dict[str, Any], stage2: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        guidance = (
            '- Choose "Nonlocal info" only if meaning depends on external module/service/config/deploy agreement.\n'
            '- If explanation is local intent for this code location, choose "Not a smell".'
        )
        return self._pair_check(sample, stage1, stage2, NOT_SMELL, "Nonlocal info", guidance)

    def _pair_check_not_smell_vs_too_much_info(self, sample: Dict[str, str], stage1: Dict[str, Any], stage2: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        guidance = (
            '- Choose "Too much info" only if comment contains unnecessary implementation/process detail.\n'
            '- If concise and mainly conveys why/intent/constraint, choose "Not a smell".'
        )
        return self._pair_check(sample, stage1, stage2, NOT_SMELL, "Too much info", guidance)

    def _pair_check_not_smell_vs_beautification(self, sample: Dict[str, str], stage1: Dict[str, Any], stage2: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        guidance = (
            '- Choose "Beautification" only for pure headers/separators with almost no behavioral or semantic content.\n'
            '- If the comment contains purpose, condition, effect, or concrete behavior, choose "Not a smell".\n'
            '- Do not choose "Beautification" just because the comment is short or formatted like a title.'
        )
        return self._pair_check(sample, stage1, stage2, NOT_SMELL, "Beautification", guidance)

    def _second_check(self, sample: Dict[str, str], stage1: Dict[str, Any], stage2: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        system_prompt = "You are a tie-breaker reviewer. Return JSON only."
        user_prompt = f"""
Do automatic second check for uncertain detection.
Allowed fine labels: {json.dumps(LABELS_SMELL, ensure_ascii=False)} and "Not a smell".
Return JSON with fields:
- final_label
- final_confidence
- top2_labels (2 items)
- reason_short

Review guidance:
- If comment gives rationale/intent/constraint not obvious from code, prefer "Not a smell".
- Only use "Obvious" when comment mostly restates visible code action.

Sample:
language: {sample.get('language', '')}
code:
{sample.get('code', '')}
comment:
{sample.get('comment', '')}

Previous stage1: {json.dumps(stage1, ensure_ascii=False)}
Previous stage2: {json.dumps(stage2, ensure_ascii=False) if stage2 else "null"}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=0.0, max_tokens=500)

    def _direct_multiclass_check(self, sample: Dict[str, str]) -> Dict[str, Any]:
        dynamic_examples = _build_stage2_dynamic_examples(sample)
        system_prompt = "You are a strict code-comment smell classifier. Return JSON only."
        user_prompt = f"""
Code comments are considered smells if they degrade software quality or do not help code comprehension.
Classify the inline comment using exactly one label from:
{json.dumps(ALL_LABELS, ensure_ascii=False)}

Decision rules:
- Not all short comments are smells.
- Choose "Obvious" only if the comment mostly restates visible local code action.
- Choose "Not a smell" when the comment provides useful local rationale/intent/constraint/risk and no strong smell evidence is present.
- Use "Vague" only if key object/condition/effect remains unclear.
- Use "Beautification" only for pure headers/separators with near-zero semantic content.
- Do not default to "Not a smell" for TODO/FIXME, commented-out code, contradiction with code, explicit external/system-wide dependency, or clearly underspecified comments.
- If comment only restates code without useful context, choose "Obvious".

Taxonomy:
{json.dumps(TAXONOMY, ensure_ascii=False)}

Reference examples:
{dynamic_examples}

Return JSON fields:
- label
- confidence (0..1)
- top2_labels (2 items)
- reason_short

Sample:
language: {sample.get('language', '')}
code:
{sample.get('code', '')}
comment:
{sample.get('comment', '')}
""".strip()
        return self.client.chat_json(system_prompt, user_prompt, temperature=0.0, max_tokens=550)

    def _top2_margin(self, raw_top2: Any, label: str, conf: float) -> float:
        top2 = self._normalize_top2(raw_top2, label, conf)
        if len(top2) < 2:
            return 1.0
        scores = sorted([float(top2[0]["score"]), float(top2[1]["score"])], reverse=True)
        return max(0.0, min(1.0, scores[0] - scores[1]))

    def detect_one(self, sample: Dict[str, str]) -> Tuple[DetectionOutput, Dict[str, Any], Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
        stage1 = self._stage1(sample)

        s1_label_raw = norm(stage1.get("label"))
        s1_label = canon_label(s1_label_raw)
        s1_conf = float(stage1.get("confidence", 0.5))
        s1_top2 = stage1.get("top2_labels", []) or []
        s1_reason = norm(stage1.get("reason_short"))
        s1_unexpected_label = False

        if s1_label_raw.lower() == "smell":
            s1_binary_smell = True
        elif s1_label == NOT_SMELL:
            s1_binary_smell = False
        elif s1_label in LABELS_SMELL:
            # Robust fallback: treat leaked fine-grained labels as "Smell" in Stage-1.
            s1_binary_smell = True
            s1_unexpected_label = True
        else:
            s1_binary_smell = True
            s1_unexpected_label = True

        stage2: Optional[Dict[str, Any]] = None
        second_raw: Optional[Dict[str, Any]] = None
        s1_margin = self._top2_margin(s1_top2, NOT_SMELL if not s1_binary_smell else "Smell", s1_conf)
        skip_stage2 = (
            (not s1_binary_smell)
            and s1_conf >= 0.93
            and s1_margin >= 0.25
        )

        if not skip_stage2:
            stage2 = self._stage2(sample)
            label = canon_label(norm(stage2.get("label")) or "Obvious")
            conf = float(stage2.get("confidence", s1_conf))
            top2 = stage2.get("top2_labels", []) or []
            reason = norm(stage2.get("reason_short")) or s1_reason
            is_smelly = label != NOT_SMELL
            signals = stage2.get("signals", {}) if isinstance(stage2.get("signals"), dict) else {}
        else:
            label = NOT_SMELL
            conf = s1_conf
            top2 = s1_top2 if isinstance(s1_top2, list) else []
            reason = s1_reason
            is_smelly = False
            signals = {}

        if label not in ALL_LABELS:
            label = NOT_SMELL if not is_smelly else "Obvious"

        if not isinstance(top2, list) or len(top2) < 2:
            if is_smelly:
                top2 = [{"label": label, "score": round(conf, 4)}, {"label": "Obvious", "score": round(max(0.0, 1.0 - conf), 4)}]
            else:
                top2 = [{"label": NOT_SMELL, "score": round(conf, 4)}, {"label": "Obvious", "score": round(max(0.0, 1.0 - conf), 4)}]

        # Hybrid verification inspired by direct notebook-style classification:
        # only use it for the hardest Not-a-smell vs Obvious boundary.
        pre_hybrid_margin = self._top2_margin(top2, label, conf)
        pre_hybrid_top1, pre_hybrid_top2 = self._extract_top2_labels(top2, label, conf)
        pre_hybrid_pair_set = {pre_hybrid_top1, pre_hybrid_top2}
        if self.hybrid_verify and (
            pre_hybrid_pair_set == {NOT_SMELL, "Obvious"}
            and (conf < 0.85 or pre_hybrid_margin < 0.10)
        ):
            try:
                direct_raw = self._direct_multiclass_check(sample)
                direct_label = canon_label(norm(direct_raw.get("label")))
                direct_conf = float(direct_raw.get("confidence", conf))
                if direct_label in ALL_LABELS:
                    if direct_label == label:
                        conf = max(conf, min(1.0, direct_conf))
                    else:
                        guidance = (
                            f'- Resolve only between "{label}" and "{direct_label}".\n'
                            '- Prefer labels supported by concrete code-comment alignment.\n'
                            '- Do not punish concise but useful comments.\n'
                            '- If one option mainly restates local code and the other provides purpose/intent, prefer the latter.'
                        )
                        arb = self._pair_check(sample, stage1, stage2, label, direct_label, guidance)
                        arb_label = canon_label(norm(arb.get("final_label") or arb.get("label")))
                        arb_conf = float(arb.get("final_confidence", arb.get("confidence", conf)))
                        arb_top2 = arb.get("top2_labels", top2)
                        arb_reason = norm(arb.get("reason_short")) or reason
                        if arb_label in {label, direct_label}:
                            label = arb_label
                            conf = arb_conf
                            top2 = arb_top2 if isinstance(arb_top2, list) else top2
                            reason = arb_reason
                            is_smelly = label != NOT_SMELL
            except Exception:
                pass

        margin = self._top2_margin(top2, label, conf)
        top1_label, top2_label = self._extract_top2_labels(top2, label, conf)
        pair_set = {top1_label, top2_label}
        signal_conflict = self._signal_conflict(label, signals)
        force_not_smell_vague_pair = (
            label == NOT_SMELL
            and "Vague" in pair_set
            and conf <= 0.95
        )
        force_nonlocal_pair = bool(signals.get("references_external_module_or_config", False))
        force_toomuch_pair = bool(signals.get("contains_excess_implementation_detail", False))
        forced_pair: Optional[str] = None
        if force_nonlocal_pair and not force_toomuch_pair:
            forced_pair = "not_smell_vs_nonlocal"
        elif force_toomuch_pair and not force_nonlocal_pair:
            forced_pair = "not_smell_vs_toomuch"
        elif force_nonlocal_pair and force_toomuch_pair:
            if label == "Nonlocal info":
                forced_pair = "not_smell_vs_nonlocal"
            elif label == "Too much info":
                forced_pair = "not_smell_vs_toomuch"
            else:
                forced_pair = "not_smell_vs_nonlocal"
        borderline_threshold = max(self.second_check_threshold, self.borderline_threshold)
        should_second_check = (
            conf < self.second_check_threshold
            or s1_unexpected_label
            or margin < self.uncertainty_margin_threshold
            or signal_conflict
            or forced_pair is not None
            or force_not_smell_vague_pair
            or (label in {NOT_SMELL, "Obvious", "Vague", "Nonlocal info"} and conf < borderline_threshold)
        )
        review_reasons: List[str] = []

        if conf < self.second_check_threshold:
            review_reasons.append("low_confidence")

        if s1_unexpected_label:
            review_reasons.append("stage1_unexpected_label")

        if margin < self.uncertainty_margin_threshold:
            review_reasons.append("small_top2_margin")

        if signal_conflict:
            review_reasons.append("signal_conflict")

        if forced_pair is not None:
            review_reasons.append(forced_pair)

        if force_not_smell_vague_pair:
            review_reasons.append("not_smell_vs_vague_risk")

        if label in {NOT_SMELL, "Obvious", "Vague", "Nonlocal info"} and conf < borderline_threshold:
            review_reasons.append("borderline_label_low_confidence")

        if should_second_check:
            if forced_pair == "not_smell_vs_nonlocal":
                second_raw = self._pair_check_not_smell_vs_nonlocal(sample, stage1, stage2)
            elif forced_pair == "not_smell_vs_toomuch":
                second_raw = self._pair_check_not_smell_vs_too_much_info(sample, stage1, stage2)
            elif pair_set == {NOT_SMELL, "Obvious"}:
                second_raw = self._pair_check_not_smell_vs_obvious(sample, stage1, stage2)
            elif pair_set == {NOT_SMELL, "Vague"}:
                second_raw = self._pair_check_not_smell_vs_vague(sample, stage1, stage2)
            elif pair_set == {NOT_SMELL, "Nonlocal info"}:
                second_raw = self._pair_check_not_smell_vs_nonlocal(sample, stage1, stage2)
            elif pair_set == {NOT_SMELL, "Too much info"}:
                second_raw = self._pair_check_not_smell_vs_too_much_info(sample, stage1, stage2)
            elif pair_set == {NOT_SMELL, "Beautification"}:
                second_raw = self._pair_check_not_smell_vs_beautification(sample, stage1, stage2)
            else:
                second_raw = self._second_check(sample, stage1, stage2)

            before_second_label = label
            sec_label = canon_label(norm(second_raw.get("final_label") or second_raw.get("label") or label))
            sec_conf = float(second_raw.get("final_confidence", second_raw.get("confidence", conf)))
            sec_top2 = second_raw.get("top2_labels", top2)
            sec_reason = norm(second_raw.get("reason_short")) or reason
            if sec_label in ALL_LABELS:
                label = sec_label
                conf = sec_conf
                top2 = sec_top2 if isinstance(sec_top2, list) else top2
                reason = sec_reason
                is_smelly = label != NOT_SMELL
                if label != before_second_label:
                    conf *= 0.75

        if signal_conflict:
            conf *= 0.85
        if margin < 0.08:
            conf *= 0.90

        # route_to only means the next pipeline stage.
        # It should not be used to record whether second check was triggered.
        route_to = "repair" if label != NOT_SMELL else "evaluation"

        output = DetectionOutput(
            sample_id=str(sample.get("sample_id", "")),
            is_smelly=bool(is_smelly),
            label=label,
            confidence=round(max(0.0, min(1.0, conf)), 4),
            top2_labels=self._normalize_top2(top2, label, conf),
            reason_short=reason[:220],
            action_hint=ACTION_HINTS.get(label, "no_change"),
            route_to=route_to,
            model_source=self.client.model_source,
            second_check_used=bool(should_second_check),
            review_reason=";".join(review_reasons),
        )
        return output, stage1, stage2, second_raw

    def _normalize_top2(self, raw: Any, label: str, conf: float) -> List[Dict[str, float]]:
        items: List[Dict[str, float]] = []
        if isinstance(raw, list):
            for item in raw:
                if not isinstance(item, dict):
                    continue
                l = canon_label(norm(item.get("label")))
                if l not in ALL_LABELS:
                    continue
                s = float(item.get("score", 0.0))
                items.append({"label": l, "score": round(max(0.0, min(1.0, s)), 4)})
        if not items:
            items = [{"label": label, "score": round(conf, 4)}]
        if len(items) == 1:
            backup = NOT_SMELL if items[0]["label"] != NOT_SMELL else "Obvious"
            items.append({"label": backup, "score": round(max(0.0, 1.0 - items[0]["score"]), 4)})
        return items[:2]


def write_detection_outputs(results: List[DetectionOutput], jsonl_path: Path, csv_path: Path) -> None:
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    with jsonl_path.open("w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(asdict(r), ensure_ascii=False) + "\n")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "sample_id",
            "is_smelly",
            "label",
            "confidence",
            "top1_label",
            "top1_score",
            "top2_label",
            "top2_score",
            "reason_short",
            "action_hint",
            "route_to",
            "second_check_used",
            "review_reason",
            "model_source",
        ])
        for r in results:
            t1 = r.top2_labels[0] if r.top2_labels else {"label": "", "score": 0.0}
            t2 = r.top2_labels[1] if len(r.top2_labels) > 1 else {"label": "", "score": 0.0}
            w.writerow([
                r.sample_id,
                r.is_smelly,
                r.label,
                r.confidence,
                t1["label"],
                t1["score"],
                t2["label"],
                t2["score"],
                r.reason_short,
                r.action_hint,
                r.route_to,
                r.second_check_used,
                r.review_reason,
                r.model_source,
            ])


def write_trace(traces: List[DetectionTrace], trace_path: Path) -> None:
    trace_path.parent.mkdir(parents=True, exist_ok=True)
    with trace_path.open("w", encoding="utf-8") as f:
        for t in traces:
            payload = {
                "sample_id": t.sample_id,
                "language": t.language,
                "code": t.code,
                "comment": t.comment,
                "gold_label": t.gold_label,
                "stage1_raw": t.stage1_raw,
                "stage2_raw": t.stage2_raw,
                "second_check_raw": t.second_check_raw,
                "detection": asdict(t.detection),
            }
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def append_detection_output(result: DetectionOutput, jsonl_path: Path) -> None:
    jsonl_path.parent.mkdir(parents=True, exist_ok=True)
    with jsonl_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(asdict(result), ensure_ascii=False) + "\n")


def append_trace(trace: DetectionTrace, trace_path: Path) -> None:
    trace_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "sample_id": trace.sample_id,
        "language": trace.language,
        "code": trace.code,
        "comment": trace.comment,
        "gold_label": trace.gold_label,
        "stage1_raw": trace.stage1_raw,
        "stage2_raw": trace.stage2_raw,
        "second_check_raw": trace.second_check_raw,
        "detection": asdict(trace.detection),
    }
    with trace_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def load_existing_detection_outputs(jsonl_path: Path) -> List[DetectionOutput]:
    if not jsonl_path.exists():
        return []
    results: List[DetectionOutput] = []
    output_fields = {field.name for field in fields(DetectionOutput)}
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                payload = json.loads(line)
                results.append(DetectionOutput(**{k: v for k, v in payload.items() if k in output_fields}))
    return results


def load_existing_detection_traces(trace_path: Path) -> List[DetectionTrace]:
    if not trace_path.exists():
        return []
    traces: List[DetectionTrace] = []
    output_fields = {field.name for field in fields(DetectionOutput)}
    with trace_path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            payload = json.loads(line)
            detection_payload = payload["detection"]
            detection = DetectionOutput(**{k: v for k, v in detection_payload.items() if k in output_fields})
            traces.append(
                DetectionTrace(
                    sample_id=str(payload.get("sample_id", "")),
                    language=payload.get("language", ""),
                    code=payload.get("code", ""),
                    comment=payload.get("comment", ""),
                    gold_label=payload.get("gold_label"),
                    stage1_raw=payload.get("stage1_raw", {}),
                    stage2_raw=payload.get("stage2_raw"),
                    second_check_raw=payload.get("second_check_raw"),
                    detection=detection,
                )
            )
    return traces


def run_detection(
    dataset_path: Optional[Path] = None,
    output_dir: Path = Path("outputs"),
    limit: Optional[int] = None,
    confidence_threshold: float = 0.75,
    second_check_threshold: float = 0.80,
    temperature: float = 0.0,
    borderline_threshold: float = 0.90,
    uncertainty_margin_threshold: float = 0.20,
    fail_fast: bool = True,
    progress_callback: Optional[
        Callable[[str, str, Optional[int], Optional[int], str, Optional[str], Optional[Dict[str, Any]]], None]
    ] = None,
) -> Dict[str, Path]:
    ds_path = resolve_dataset_path(dataset_path)
    if not ds_path.exists():
        raise FileNotFoundError(f"Dataset not found: {ds_path}")

    samples = load_samples_from_xlsx(ds_path, limit=limit)
    total = len(samples)
    agent = DetectionAgent(
        confidence_threshold=confidence_threshold,
        second_check_threshold=second_check_threshold,
        temperature=temperature,
        borderline_threshold=borderline_threshold,
        uncertainty_margin_threshold=uncertainty_margin_threshold,
    )

    jsonl_path = output_dir / "detection_results.jsonl"
    csv_path = output_dir / "detection_results.csv"
    trace_path = output_dir / "detection_trace.jsonl"

    results = load_existing_detection_outputs(jsonl_path)
    traces = load_existing_detection_traces(trace_path)
    processed_ids = {r.sample_id for r in results}
    if traces:
        processed_ids &= {t.sample_id for t in traces}
    if results and len(results) != len(traces):
        processed_ids = {t.sample_id for t in traces}
        results = [r for r in results if r.sample_id in processed_ids]

    if progress_callback is not None:
        progress_callback(
            "detection",
            "running",
            len(processed_ids),
            total,
            "Detection stage started",
            None,
            {"dataset": str(ds_path), "resumed": bool(processed_ids)},
        )

    for idx, s in enumerate(samples, start=1):
        sample_id = str(s.get("sample_id", ""))
        if sample_id in processed_ids:
            continue
        try:
            out, st1, st2, sec = agent.detect_one(s)
        except (LLMRequestError, ValueError, KeyError, TypeError) as exc:
            write_detection_outputs(results, jsonl_path, csv_path)
            write_trace(traces, trace_path)
            if progress_callback is not None:
                progress_callback(
                    "detection",
                    "failed",
                    idx,
                    total,
                    "Detection failed on sample",
                    str(s.get("sample_id", "")),
                    {"error": str(exc)},
                )
            if fail_fast:
                raise RuntimeError(f"Detection failed on sample_id={s.get('sample_id', '')}: {exc}") from exc
            raise

        results.append(out)
        trace = DetectionTrace(
            sample_id=str(s.get("sample_id", "")),
            language=s.get("language", ""),
            code=s.get("code", ""),
            comment=s.get("comment", ""),
            gold_label=canon_label(s.get("gold_label", "")) if s.get("gold_label") else None,
            stage1_raw=st1,
            stage2_raw=st2,
            second_check_raw=sec,
            detection=out,
        )
        traces.append(trace)
        append_detection_output(out, jsonl_path)
        append_trace(trace, trace_path)
        if progress_callback is not None:
            progress_callback(
                "detection",
                "running",
                idx,
                total,
                "Detection sample processed",
                str(s.get("sample_id", "")),
                {"label": out.label, "confidence": out.confidence},
            )

    write_detection_outputs(results, jsonl_path, csv_path)
    write_trace(traces, trace_path)
    if progress_callback is not None:
        progress_callback(
            "detection",
            "completed",
            total,
            total,
            "Detection stage completed",
            None,
            {"results_jsonl": str(jsonl_path), "trace_jsonl": str(trace_path)},
        )

    return {
        "detection_results_jsonl": jsonl_path,
        "detection_results_csv": csv_path,
        "detection_trace_jsonl": trace_path,
    }


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Run Detection Agent on Data/Dataset.xlsx")
    p.add_argument("--dataset", type=Path, default=None, help="Path to Dataset.xlsx")
    p.add_argument("--output-dir", type=Path, default=Path("outputs"))
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--confidence-threshold", type=float, default=0.75)
    p.add_argument("--second-check-threshold", type=float, default=0.80)
    p.add_argument("--borderline-threshold", type=float, default=0.90)
    p.add_argument("--uncertainty-margin-threshold", type=float, default=0.20)
    p.add_argument("--temperature", type=float, default=0.0)
    return p.parse_args()


def main() -> None:
    args = parse_args()

    run_detection(
        dataset_path=args.dataset,
        output_dir=args.output_dir,
        limit=args.limit,
        confidence_threshold=args.confidence_threshold,
        second_check_threshold=args.second_check_threshold,
        temperature=args.temperature,
        borderline_threshold=args.borderline_threshold,
        uncertainty_margin_threshold=args.uncertainty_margin_threshold,
    )


if __name__ == "__main__":
    main()
